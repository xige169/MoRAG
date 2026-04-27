from openpyxl import load_workbook


def parse_xlsx(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    blocks = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append("\t".join(cells))
        if rows:
            blocks.append({
                "content": f"[Sheet: {sheet.title}]\n" + "\n".join(rows),
                "page_number": None,
            })
    wb.close()
    return blocks
